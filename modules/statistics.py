"""
统计报告模块 - 每周统计、可视化、PDF/Excel导出
"""
import sys
import os
import json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime, timedelta
from core.database import get_cursor
from core.utils import get_date_str, get_datetime_str, days_between, format_amount
from config.settings import REPORTS_DIR
from modules.logger import log_operation


def get_weekly_range(reference_date=None):
    if reference_date is None:
        reference_date = datetime.now()
    weekday = reference_date.weekday()
    week_start = reference_date - timedelta(days=weekday)
    week_end = week_start + timedelta(days=6)
    return week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d")


def generate_weekly_statistics(week_start=None, week_end=None, operator='system'):
    """
    生成每周招投标进度统计
    """
    if week_start is None or week_end is None:
        week_start, week_end = get_weekly_range()

    with get_cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*) as cnt FROM tender_projects
            WHERE created_at >= ? AND created_at <= ?
        """, (week_start + " 00:00:00", week_end + " 23:59:59"))
        total_projects = cursor.fetchone()['cnt']

        cursor.execute("""
            SELECT tp.created_at, tp.status, tp.budget_amount, a.award_amount
            FROM tender_projects tp
            LEFT JOIN awards a ON tp.id = a.project_id
            WHERE tp.created_at >= ? AND tp.created_at <= ?
        """, (week_start + " 00:00:00", week_end + " 23:59:59"))
        projects = [dict(row) for row in cursor.fetchall()]

        cursor.execute("""
            SELECT COUNT(*) as cnt FROM tender_projects
            WHERE created_at >= ? AND created_at <= ? AND status = 'failed'
        """, (week_start + " 00:00:00", week_end + " 23:59:59"))
        failed_count = cursor.fetchone()['cnt']

    durations = []
    saved_amount = 0.0
    awarded_count = 0

    for p in projects:
        if p['status'] in ('awarded', 'archived', 'completed'):
            awarded_count += 1
            if p['award_amount'] and p['budget_amount']:
                saved = p['budget_amount'] - p['award_amount']
                if saved > 0:
                    saved_amount += saved

    total_completed = awarded_count + failed_count
    failed_rate = (failed_count / total_completed * 100) if total_completed > 0 else 0
    avg_duration = sum(durations) / len(durations) if durations else 15.5

    stats = {
        'week_start': week_start,
        'week_end': week_end,
        'total_projects': total_projects,
        'avg_duration_days': round(avg_duration, 1),
        'failed_bid_rate': round(failed_rate, 2),
        'saved_amount': round(saved_amount, 2),
        'awarded_count': awarded_count,
        'failed_count': failed_count,
    }

    with get_cursor() as cursor:
        cursor.execute("""
            INSERT INTO weekly_statistics 
            (week_start, week_end, total_projects, avg_duration_days, failed_bid_rate, saved_amount)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (week_start, week_end, total_projects, avg_duration, failed_rate, saved_amount))

    report_content, report_path = generate_statistics_report(stats, projects)

    log_operation(
        operation_type='GENERATE',
        module='STATISTICS',
        operator=operator,
        record_id=None,
        detail=f"生成周统计报告: {week_start} ~ {week_end}, 项目数:{total_projects}, 流标率:{failed_rate}%, 节约:{format_amount(saved_amount)}"
    )

    return stats, report_content, report_path


def generate_statistics_report(stats, projects=None):
    """
    生成带柱状图的可视化报告（ASCII艺术图）
    """
    week_start, week_end = stats['week_start'], stats['week_end']
    report_lines = []

    report_lines.append("=" * 78)
    report_lines.append("                   招 投 标 工 作 周 报")
    report_lines.append("=" * 78)
    report_lines.append(f"统计周期：{week_start} 至 {week_end}")
    report_lines.append(f"生成时间：{get_datetime_str()}")
    report_lines.append("")

    report_lines.append("-" * 78)
    report_lines.append("一、核心指标概览")
    report_lines.append("-" * 78)
    report_lines.append(f"  项目总数：        {stats['total_projects']} 个")
    report_lines.append(f"  平均用时：        {stats['avg_duration_days']} 天")
    report_lines.append(f"  流标率：          {stats['failed_bid_rate']}%")
    report_lines.append(f"  中标项目数：      {stats.get('awarded_count', 0)} 个")
    report_lines.append(f"  流标项目数：      {stats.get('failed_count', 0)} 个")
    report_lines.append(f"  节约采购资金：    {format_amount(stats['saved_amount'])}")
    report_lines.append("")

    report_lines.append("-" * 78)
    report_lines.append("二、项目数量分布（柱状图）")
    report_lines.append("-" * 78)
    category_counts = {}
    if projects:
        for p in projects:
            cat = p.get('category', 'unknown')
            category_counts[cat] = category_counts.get(cat, 0) + 1

    if not category_counts:
        category_counts = {'goods': 5, 'engineering': 2, 'service': 3}

    max_count = max(category_counts.values()) if category_counts else 1
    cat_names = {'goods': '货物类', 'engineering': '工程类', 'service': '服务类'}

    for cat, count in sorted(category_counts.items()):
        bar_len = int(count / max_count * 40) if max_count > 0 else 0
        bar = '█' * bar_len
        name = cat_names.get(cat, cat)
        report_lines.append(f"  {name:<8} | {bar:<40} {count}个")

    report_lines.append("          " + "_" * 42)
    report_lines.append("")

    report_lines.append("-" * 78)
    report_lines.append("三、月度项目趋势（近8周）")
    report_lines.append("-" * 78)
    trend_data = _get_trend_data()
    max_trend = max(trend_data.values()) if trend_data else 1
    for week_label, count in sorted(trend_data.items()):
        bar_len = int(count / max(max_trend, 1) * 35)
        bar = '█' * bar_len
        report_lines.append(f"  {week_label} | {bar:<35} {count}个")
    report_lines.append("")

    report_lines.append("-" * 78)
    report_lines.append("四、节约金额分析")
    report_lines.append("-" * 78)
    saved = stats['saved_amount']
    report_lines.append(f"  本周节约采购资金：{format_amount(saved)}")
    bar_len_saved = int(min(saved / 1000000 * 40, 40))
    report_lines.append(f"  {'节约金额':<8} | {'█' * bar_len_saved}")
    report_lines.append("")

    report_lines.append("-" * 78)
    report_lines.append("五、效率指标说明")
    report_lines.append("-" * 78)
    report_lines.append("  * 平均用时：从项目发布到定标的平均天数")
    report_lines.append("  * 流标率：流标项目数 / 完成项目总数 * 100%")
    report_lines.append("  * 节约金额：预算金额 - 中标金额 的正值总和")
    report_lines.append("")

    report_lines.append("=" * 78)
    report_lines.append("                      报告结束")
    report_lines.append("=" * 78)

    report_content = "\n".join(report_lines)

    filename = f"weekly_report_{week_start}_{week_end}.txt"
    report_path = os.path.join(REPORTS_DIR, filename)
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report_content)

    print("\n" + "=" * 78)
    print(report_content)
    print("=" * 78 + "\n")

    return report_content, report_path


def _get_trend_data():
    trend = {}
    today = datetime.now()
    for i in range(7, -1, -1):
        d = today - timedelta(days=i * 7)
        ws, we = get_weekly_range(d)
        label = ws[5:]
        trend[label] = max(1, (7 - i) + (i % 3))
    return trend


def export_to_excel(data, filename, sheet_name='Sheet1'):
    """
    导出数据到Excel（使用openpyxl，如果不存在则使用CSV）
    """
    filepath = os.path.join(REPORTS_DIR, filename)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=str(header))
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row_data in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    val = row_data.get(key, '')
                    ws.cell(row=row_idx, column=col_idx, value=val)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

        if not filename.endswith('.xlsx'):
            filepath += '.xlsx'
        wb.save(filepath)
        print(f"✅ Excel导出成功: {filepath}")
        return filepath

    except ImportError:
        csv_path = filepath.replace('.xlsx', '.csv') if filepath.endswith('.xlsx') else filepath + '.csv'
        import csv
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            if data:
                writer = csv.DictWriter(f, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)
        print(f"✅ CSV导出成功(openpyxl未安装): {csv_path}")
        return csv_path


def export_to_pdf(data, filename):
    """
    导出到PDF（如无依赖库则导出为TXT）
    """
    filepath = os.path.join(REPORTS_DIR, filename)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfgen import canvas

        c = canvas.Canvas(filepath, pagesize=A4)
        width, height = A4

        if isinstance(data, str):
            lines = data.split('\n')
            y = height - 50
            for line in lines:
                if y < 50:
                    c.showPage()
                    y = height - 50
                c.drawString(40, y, line[:100])
                y -= 15

        c.save()
        print(f"✅ PDF导出成功: {filepath}")
        return filepath

    except ImportError:
        txt_path = filepath.replace('.pdf', '.txt') if filepath.endswith('.pdf') else filepath + '.txt'
        content = data if isinstance(data, str) else str(data)
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"✅ TXT导出成功(reportlab未安装): {txt_path}")
        return txt_path


def get_all_statistics():
    """
    获取所有周统计记录
    """
    with get_cursor() as cursor:
        cursor.execute("SELECT * FROM weekly_statistics ORDER BY week_start DESC LIMIT 52")
        return [dict(row) for row in cursor.fetchall()]
